import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path

# The script establishes dynamic file paths to prevent FileNotFoundError when loading
script_dir = Path(__file__).parent
filename = script_dir / 'transactions.xlsx'

def transactions(file_path):
    # The workbook is loaded into memory and the specific sheet is selected
    wb = xl.load_workbook(file_path)
    sheet = wb['Sheet1'] # Sheet names are case-sensitive

    # The top-left cell is targeted for structural verification
    cell = sheet['a1']
    cell = sheet.cell(1, 1) # Targets the transaction_id header

    # Terminal outputs are generated to confirm data boundaries
    print(cell.value) 
    print(sheet.max_row) # Outputs the total number of populated rows
    print(sheet.max_column) # Outputs the total number of populated columns

    # The iterative process begins at row 2 to bypass the column headers
    for row in range(2, sheet.max_row + 1): 
        cell = sheet.cell(row, 3) # Extracts data from the prices column
        
        # The cell value is explicitly cast to a string before the currency symbol is stripped
        num_price = str(cell.value).replace("£", "")
        
        # The cleaned string is cast to a float and a 10% discount is mathematically applied
        discount = float(num_price) * 0.9
        
        # The resulting discounted value is written into the adjacent fourth column
        discount_section = sheet.cell(row, 4) 
        discount_section.value = discount

    # A data reference block is generated targeting the newly populated discount column
    values = Reference(
        sheet, 
        min_row=2, 
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )

    # A bar chart object is instantiated, fed the reference data, and anchored to cell E2
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # The modified workbook is saved as a new file version to preserve the original dataset
    wb.save('transactionv1.xlsx')


# The execution block triggers the pipeline using the dynamic path
if __name__ == "__main__":
    transactions(filename)